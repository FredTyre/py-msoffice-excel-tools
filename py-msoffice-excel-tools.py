import csv
import openpyxl
from shutil import copyfile

input_directory = "input/"
output_directory = "output/"

def convert_csv_to_xlsx(filename_without_extension):
    csv_file_path = input_directory + filename_without_extension + '.csv'
    msexcel_file_path = filename_without_extension + '.xlsx'
    output_msexcel_file_path = output_directory + msexcel_file_path
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    with open(csv_file_path) as file_handler:
        reader = csv.reader(file_handler, delimiter=',')
        for row in reader:
            worksheet.append(row)

    workbook.save(msexcel_output_file_path)
    copyfile(msexcel_output_file_path, input_directory + msexcel_file_path)

def rows_to_cols(filename_without_extension, area_columns, area_rows):
    filename = filename_without_extension + '.xlsx'
    
    workbook = openpyxl.load_workbook(input_directory + filename)
    worksheet = workbook.active

    area_read_in = []
    for current_column in range(1, area_columns + 1):
        for current_row in range(1, area_rows + 1):
            print(worksheet.cell(row = current_row, column = current_column).value)
            area_read_in.append(worksheet.cell(row = current_row, column = current_column).value)

    print(len(area_read_in))
    print(area_read_in)

    area_read_in_index = 0
    for current_row in range(1, area_rows + 1):
        for current_column in range(1, area_columns + 1):
            worksheet.cell(row = current_row, column = current_column).value = area_read_in[area_read_in_index]
            area_read_in_index += 1
    
    workbook.save(output_directory + filename)
    
#convert_csv_to_xlsx('RowsColumnsInversed')
rows_to_cols('RowsColumnsInversed', 5, 5)
